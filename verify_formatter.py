"""End-to-end verification against reference output."""
from openpyxl import load_workbook

from formatter import format_shipment

INPUT = r"c:\Users\Administrator\Downloads\FBA19FJ35VLB.xlsx"
OUTPUT = r"c:\Users\Administrator\Downloads\FBA19FJ35VLB-Box Contents.xlsx"
PL_INPUT = r"c:\Users\Administrator\Downloads\FBA19FJ1KRVH-PL.xls"


def read_cells(path_or_buf):
    wb = load_workbook(path_or_buf, data_only=False)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        result[name] = [
            [(c.value, type(c.value).__name__) for c in row if c.value is not None]
            for row in ws.iter_rows(values_only=False)
        ]
    return wb.sheetnames, result


with open(INPUT, "rb") as f:
    buf, filename, summary = format_shipment(f, INPUT)

assert filename == "FBA19FJ35VLB-Box Contents.xlsx"

gen_names, gen_data = read_cells(buf)
ref_names, ref_data = read_cells(OUTPUT)

assert gen_names == ref_names == ["Box Contents", "Weights and Dimensions"]

for sheet in ref_names:
    assert gen_data[sheet] == ref_data[sheet], f"Mismatch in {sheet}"

print("All checks passed.")
print(f"Filename: {filename}")
print(f"Summary: {summary}")

with open(PL_INPUT, "rb") as f:
    pl_buf, pl_filename, pl_summary = format_shipment(f, PL_INPUT)

assert pl_filename == "FBA19FJ1KRVH-Box Contents.xlsx"
assert pl_summary == {
    "shipment_id": "FBA19FJ1KRVH",
    "carton_count": 25,
    "line_count": 47,
    "total_units": 115,
    "total_weight": 344,
}

pl_wb = load_workbook(pl_buf, data_only=True)
pl_contents = pl_wb["Box Contents"]
assert pl_contents["A2"].value == 1
assert pl_contents["B2"].value == "197642129593"
assert pl_contents["C2"].value == 1

pl_dims = pl_wb["Weights and Dimensions"]
assert pl_dims["A2"].value == 1
assert pl_dims["B2"].value == 14
assert pl_dims["C2"].value == 24
assert pl_dims["D2"].value == 16
assert pl_dims["E2"].value == 11

print("PL format checks passed.")
print(f"PL filename: {pl_filename}")
print(f"PL summary: {pl_summary}")
