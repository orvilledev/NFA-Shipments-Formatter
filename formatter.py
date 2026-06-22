import csv
import io
import os
import re
from io import BytesIO

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

SUPPORTED_EXTENSIONS = {
    ".xlsx",
    ".xlsm",
    ".xltx",
    ".xltm",
    ".xls",
    ".csv",
    ".tsv",
}


def parse_input(source, filename: str | None = None) -> tuple[str, list[dict]]:
    """Parse an FBA Carton Detail file into structured carton data."""
    rows = _load_rows(source, filename)

    shipment_id = _extract_shipment_id(rows, filename)
    cartons: list[dict] = []
    current = None

    for row in rows:
        if _is_carton_row(row):
            if current:
                cartons.append(current)
            current = {
                "carton_id": _format_identifier(row[1]),
                "items": [],
                "weight": None,
                "length": None,
                "width": None,
                "height": None,
            }
            if _has_dims(row):
                _apply_dims(current, row)
        elif current and _is_total_row(row):
            if _has_dims(row):
                _apply_dims(current, row)
        elif current and _is_product_row(row):
            current["items"].append({"upc": _format_upc(row[1]), "qty": int(row[3])})

    if current:
        cartons.append(current)

    if not shipment_id:
        raise ValueError(
            "Could not find shipment ID. Expected an FBA shipment ID near the top of the file "
            "or in the uploaded filename."
        )
    if not cartons:
        raise ValueError("No cartons found. Expected rows starting with 'Carton#:'.")

    return shipment_id, cartons


def build_output(shipment_id: str, cartons: list[dict]) -> tuple[BytesIO, str]:
    """Build the formatted Box Contents workbook."""
    wb = Workbook()

    ws_contents = wb.active
    ws_contents.title = "Box Contents"
    ws_contents.append(["Box Number", "UPC", "Qty"])

    for box_number, carton in enumerate(cartons, 1):
        for item in carton["items"]:
            ws_contents.append([box_number, item["upc"], item["qty"]])

    for row_idx in range(2, ws_contents.max_row + 1):
        ws_contents.cell(row=row_idx, column=2).number_format = "@"

    _auto_fit_columns(ws_contents, min_widths={"B": 14})

    ws_dims = wb.create_sheet("Weights and Dimensions")
    ws_dims.append(
        ["Box Number", "Weight", "Carton Length", "Carton Width", "Carton Height"]
    )
    for box_number, carton in enumerate(cartons, 1):
        ws_dims.append(
            [
                box_number,
                carton["weight"],
                carton["length"],
                carton["width"],
                carton["height"],
            ]
        )

    _auto_fit_columns(ws_dims)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{shipment_id}-Box Contents.xlsx"
    return buffer, filename


def format_shipment(source, filename: str | None = None) -> tuple[BytesIO, str, dict]:
    """Parse input and return output bytes, filename, and summary stats."""
    if filename is None and hasattr(source, "name"):
        filename = source.name
    shipment_id, cartons = parse_input(source, filename)
    output_buffer, filename = build_output(shipment_id, cartons)

    total_units = sum(item["qty"] for carton in cartons for item in carton["items"])
    total_weight = sum(carton["weight"] or 0 for carton in cartons)

    summary = {
        "shipment_id": shipment_id,
        "carton_count": len(cartons),
        "line_count": sum(len(c["items"]) for c in cartons),
        "total_units": total_units,
        "total_weight": total_weight,
    }
    return output_buffer, filename, summary


def _read_bytes(source) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, str):
        with open(source, "rb") as handle:
            return handle.read()
    if hasattr(source, "read"):
        data = source.read()
        if hasattr(source, "seek"):
            source.seek(0)
        return data
    raise TypeError("Unsupported input source type.")


def _get_extension(filename: str | None) -> str:
    if not filename:
        return ".xlsx"
    return os.path.splitext(filename)[1].lower()


def _load_rows(source, filename: str | None = None) -> list[tuple]:
    data = _read_bytes(source)
    ext = _get_extension(filename)

    if ext not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported file type '{ext or '(none)'}'. Supported: {supported}")

    if ext in {".csv", ".tsv"}:
        return _load_rows_csv(data, delimiter="\t" if ext == ".tsv" else None)
    if ext == ".xls":
        return _load_rows_xls(data)
    return _load_rows_xlsx(BytesIO(data))


def _load_rows_xlsx(stream) -> list[tuple]:
    wb = load_workbook(stream, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(_pad_row(row))
    return rows


def _load_rows_xls(data: bytes) -> list[tuple]:
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        row = [
            _normalize_cell(sheet.cell_value(row_idx, col_idx))
            if col_idx < sheet.ncols
            else None
            for col_idx in range(max(sheet.ncols, 9))
        ]
        rows.append(_pad_row(row))
    return rows


def _load_rows_csv(data: bytes, delimiter: str | None = None) -> list[tuple]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode CSV file. Save it as UTF-8 and try again.")

    if delimiter is None:
        try:
            delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
        except csv.Error:
            delimiter = ","

    rows = []
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        if not any(cell.strip() for cell in row):
            continue
        rows.append(_pad_row(_normalize_cell(cell) or None for cell in row))
    return rows


def _pad_row(row) -> tuple:
    values = [_normalize_cell(value) for value in row]
    while len(values) < 9:
        values.append(None)
    return tuple(values[:9])


def _normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    return value


def _extract_shipment_id(rows: list[tuple], filename: str | None) -> str | None:
    for row in rows:
        if row[0] == "FBA Carton Detail" and row[1]:
            return str(row[1]).strip()

    for row in rows[:5]:
        candidate = _normalize_label(row[0])
        if candidate and re.fullmatch(r"FBA[A-Z0-9]+", candidate, re.IGNORECASE):
            return candidate.upper()

    if filename:
        match = re.search(r"(FBA[A-Z0-9]+)", os.path.basename(filename), re.IGNORECASE)
        if match:
            return match.group(1).upper()

    return None


def _is_carton_row(row: tuple) -> bool:
    label = _normalize_label(row[0])
    return bool(label and (label == "Carton#:" or label.startswith("Carton#:")))


def _is_total_row(row: tuple) -> bool:
    for index in (0, 2):
        if _normalize_label(row[index]) == "Total":
            return True
    return False


def _is_footer_row(row: tuple) -> bool:
    label = _normalize_label(row[0])
    return bool(label and label.startswith("Total Ctns:"))


def _is_header_row(row: tuple) -> bool:
    return _normalize_label(row[0]) == "Sku"


def _is_product_row(row: tuple) -> bool:
    if _is_header_row(row) or _is_carton_row(row) or _is_total_row(row) or _is_footer_row(row):
        return False
    if row[0] is None or row[3] is None:
        return False
    if row[1] is None:
        return False
    try:
        int(row[3])
    except (TypeError, ValueError):
        return False
    return True


def _has_dims(row: tuple) -> bool:
    return any(row[index] is not None for index in (5, 6, 7, 8))


def _apply_dims(carton: dict, row: tuple) -> None:
    if row[5] is not None:
        carton["weight"] = _as_int(row[5])
    if row[6] is not None:
        carton["length"] = _as_float(row[6])
    if row[7] is not None:
        carton["width"] = _as_float(row[7])
    if row[8] is not None:
        carton["height"] = _as_float(row[8])


def _normalize_label(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _format_identifier(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _format_upc(value) -> str:
    if value is None:
        raise ValueError("Missing UPC value in input file.")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _auto_fit_columns(ws, min_widths: dict[str, float] | None = None) -> None:
    """Set column widths from content so values are visible when the file opens."""
    min_widths = min_widths or {}
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        width = max(max_length + 2, min_widths.get(column_letter, 0))
        ws.column_dimensions[column_letter].width = width


def _as_int(value):
    if value is None:
        return None
    return int(value)


def _as_float(value):
    if value is None:
        return None
    return float(value)
